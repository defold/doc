import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def create_workbook():
    """
    Create Excel workbook and set basic formatting
    
    Returns:
        tuple: (workbook object, worksheet object)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docs Structure Comparison"
    
    # Set headers
    headers = ["File Path", "File Extension", "Top Level Directory", "Source Version Exists", "Target Version Exists", 
               "Status", "Source Version File Size(KB)", "Target Version File Size(KB)", 
               "Markdown Syntax Consistency", "Formula"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    return wb, ws


def get_status_fills():
    """
    Get fill colors for different statuses
    
    Returns:
        dict: Dictionary containing fill colors for different statuses
    """
    return {
        'consistent': PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid'),  # Green
        'source_only': PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid'),  # Red
        'target_only': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')   # Yellow
    }


def write_file_info(ws, row_num, file_path, file_info, source_exists, target_exists, status_fills):
    """
    Write file information to worksheet
    
    Parameters:
        ws: Worksheet object
        row_num: Row number
        file_path: File path
        file_info: File information dictionary
        source_exists: Whether source version exists
        target_exists: Whether target version exists
        status_fills: Status fill color dictionary
        
    Returns:
        str: File status
    """
    # Determine status
    if source_exists and target_exists:
        status = "Consistent"
        status_fill = status_fills['consistent']
    elif source_exists and not target_exists:
        status = "Source Only"
        status_fill = status_fills['source_only']
    elif not source_exists and target_exists:
        status = "Target Only"
        status_fill = status_fills['target_only']
    else:
        status = "Does Not Exist"
        status_fill = None
    
    # Write basic information
    ws.cell(row=row_num, column=1, value=file_path)
    ws.cell(row=row_num, column=2, value=file_info['extension'])
    ws.cell(row=row_num, column=3, value=file_info['top_dir'])
    ws.cell(row=row_num, column=4, value="Yes" if source_exists else "No")
    ws.cell(row=row_num, column=5, value="Yes" if target_exists else "No")
    
    # Set status cell
    status_cell = ws.cell(row=row_num, column=6, value=status)
    if status_fill:
        status_cell.fill = status_fill
    
    return status


def write_file_sizes(ws, row_num, source_file_info, target_file_info):
    """
    Write file size information
    
    Parameters:
        ws: Worksheet object
        row_num: Row number
        source_file_info: Source version file information
        target_file_info: Target version file information
    """
    ws.cell(row=row_num, column=7, value=source_file_info.get('size', ''))
    ws.cell(row=row_num, column=8, value=target_file_info.get('size', ''))


def write_markdown_consistency(ws, row_num, consistency_result):
    """
    Write Markdown syntax consistency results
    
    Parameters:
        ws: Worksheet object
        row_num: Row number
        consistency_result: Consistency check result
    """
    ws.cell(row=row_num, column=9, value=consistency_result)


def write_formula_text(ws, row_num, file_path, source_dir="docs\\source", target_dir="docs\\target"):
    """
    Write formula column
    
    Parameters:
        ws: Worksheet object
        row_num: Row number
        file_path: File path
        source_dir: Source directory path
        target_dir: Target directory path
    """
    # Generate formula text based on file path
    formula_text = f"Compare {source_dir}\\{file_path} and {target_dir}\\{file_path} paragraph by paragraph with each heading as a paragraph, to ensure the target version is a complete and accurate translation of the source version."
    ws.cell(row=row_num, column=10, value=formula_text)


def adjust_column_widths(ws):
    """
    Adjust worksheet column widths
    
    Parameters:
        ws: Worksheet object
    """
    ws.column_dimensions['A'].width = 50  # File Path
    ws.column_dimensions['B'].width = 10  # File Extension
    ws.column_dimensions['C'].width = 15  # Top Level Directory
    ws.column_dimensions['D'].width = 15  # Source Version Exists
    ws.column_dimensions['E'].width = 15  # Target Version Exists
    ws.column_dimensions['F'].width = 20  # Status
    ws.column_dimensions['G'].width = 20  # Source Version File Size
    ws.column_dimensions['H'].width = 20  # Target Version File Size
    ws.column_dimensions['I'].width = 50  # Markdown Syntax Consistency
    ws.column_dimensions['J'].width = 100 # Formula


def save_workbook(wb, output_file):
    """
    Save workbook to file
    
    Parameters:
        wb: Workbook object
        output_file: Output file path
        
    Returns:
        bool: Whether save was successful
    """
    try:
        print(f"Preparing to save Excel file to: {output_file}")
        
        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"Created output directory: {output_dir}")
        
        # Save Excel file
        wb.save(output_file)
        print(f"\nCompleted! File structure comparison results have been saved to: {output_file}")
        
        # Verify if file was successfully created
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"Excel file successfully created, file size: {file_size} bytes")
            return True
        else:
            print("Error: Excel file was not successfully created")
            return False
            
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def print_statistics(ws, row_num):
    """
    Print statistics results
    
    Parameters:
        ws: Worksheet object
        row_num: Total number of rows
    """
    print(f"Compared {row_num - 2} files in total")
    
    # Statistics results
    consistent_count = sum(1 for i in range(2, row_num) if ws.cell(row=i, column=6).value == "Consistent")
    source_only_count = sum(1 for i in range(2, row_num) if ws.cell(row=i, column=6).value == "Source Only")
    target_only_count = sum(1 for i in range(2, row_num) if ws.cell(row=i, column=6).value == "Target Only")
    
    print("Statistics results:")
    print(f"- Files existing in both versions: {consistent_count}")
    print(f"- Files existing only in source version: {source_only_count}")
    print(f"- Files existing only in target version: {target_only_count}")